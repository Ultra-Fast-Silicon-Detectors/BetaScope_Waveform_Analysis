from openpyxl import Workbook, load_workbook
import configparser
import os
import json

from get_time_res import Get_Time_Resolution
from read_current import Read_Current

import logging

logging.basicConfig()
# logging.getLogger().setLevel(logging.CRITICAL)
log = logging.getLogger(__name__)

par_list = [
    "runNumber",
    "SensorName",
    "Temp",
    "Bias",
    "Resistance",
    "pulseArea",
    "pulseArea_Error",
    "Pmax",
    "Pmax_Error",
    "RMS",
    "RMS_Error",
    "Rise_Time",
    "Rise_Time_Error",
    "dvdt",
    "dvdt_Error",
    "FWHM",
    "FWHM_Error",
    "NewPulseArea",
    "NewPulseArea_Error",
    "FallTime",
    "FallTime_Error",
    "cycle",
]

INI_TO_EXCEL = {
    "runNumber": "A",
    "SensorName": "B",
    "Temp": "G",
    "Bias": "H",
    "Resistance": "L",
    "pulseArea": "J",
    "pulseArea_Error": "K",
    "Pmax": "R",
    "Pmax_Error": "S",
    "RMS": "T",
    "RMS_Error": "U",
    "Rise_Time": "Z",
    "Rise_Time_Error": "AA",
    "dvdt": "AB",
    "dvdt_Error": "AC",
    "FWHM": "AL",
    "FWHM_Error": "AM",
    "NewPulseArea": "CA",
    "NewPulseArea_Error": "CB",
    "FallTime": "DG",
    "FallTime_Error": "DH",
    "cycle": "F",
    "CFD50Time": "BW",
    "CFD50Time_Err": "BX",
    "CFD20Time": "DD",
    "CFD20Time_Err": "DE",
    "CFD50Time_Par2": "BW",
    "CFD50Time_Par2Err": "BX",
    "CFD20Time_Par2": "DD",
    "CFD20Time_Par2Err": "DE",
    "Leakage": "C",
}


def MergeExcel(ifile="_results.xlxs"):
    """
    function for read in existing excel file and merge data from ifile
    """

    # loading envir path
    output_dir = os.environ["BETASCOPE_SCRIPTS"]
    if not output_dir:
        output_dir = "/tmp/"

    if not os.path.exists(f"{output_dir}/user_data/"):
        os.makedirs(f"{output_dir}/user_data/")

    # checking merged beta result
    if os.path.exists(f"{output_dir}/user_data/merged_beta_results.xlsx"):
        no_merge_file = False
        merged_wb = load_workbook(f"{output_dir}/user_data/merged_beta_results.xlsx")
    else:
        no_merge_file = True
        merged_wb = Workbook()
        merged_wb.create_sheet("DUT")
        merged_wb.create_sheet("TRIG")

    input_wb = load_workbook(ifile)

    # for meta data
    input_dut_wp = input_wb["DUT"]
    run_number_cell = f"{INI_TO_EXCEL['runNumber']}1"
    sensor_name_cell = f"{INI_TO_EXCEL['SensorName']}1"
    input_run_number = input_dut_wp[run_number_cell].value.split("->")[0]
    input_sensor_name = input_dut_wp[sensor_name_cell].value

    input_max_row = input_wb["DUT"].max_row
    merged_max_row = merged_wb["DUT"].max_row

    # check meta data
    if os.path.exists(f"{output_dir}/user_data/merged_log.json"):
        with open(f"{output_dir}/user_data/merged_log.json") as f:
            meta_data = json.load(f)
            keyName = f"run{input_run_number}"
            if keyName in meta_data:
                start_row_in_merge = meta_data[keyName]["start"]
                end_row_in_merge = meta_data[keyName]["end"]
                diff = end_row_in_merge - start_row_in_merge
                if input_max_row > diff:
                    new_rows = input_max_row - diff
                else:
                    new_rows = 0
                end_row_in_merge += new_rows
            else:
                new_rows = 0
                start_row_in_merge = merged_max_row + 2
                end_row_in_merge = start_row_in_merge + input_max_row
            meta_data[f"run{input_run_number}"] = {
                "sensor": f"{input_sensor_name}",
                "start": start_row_in_merge,
                "end": end_row_in_merge,
            }
        with open(f"{output_dir}/user_data/merged_log.json", "w") as newf:
            json.dump(meta_data, newf)
    else:
        new_rows = 0
        with open(f"{output_dir}/user_data/merged_log.json", "w") as f:
            start_row_in_merge = merged_max_row + 1
            end_row_in_merge = start_row_in_merge + input_max_row
            meta_data = {}
            meta_data[f"run{input_run_number}"] = {
                "sensor": f"{input_sensor_name}",
                "start": start_row_in_merge,
                "end": end_row_in_merge,
            }
            json.dump(meta_data, f)

    sheets = ["DUT", "TRIG"]
    for sheet in sheets:
        merged_ws = merged_wb[sheet]
        input_ws = input_wb[sheet]
        for i in range(new_rows):
            merged_ws.insert(end_row_in_merge)
        for par in INI_TO_EXCEL.keys():
            rowCounter = start_row_in_merge
            for row in range(1, input_max_row + 1):
                input_cell = f"{INI_TO_EXCEL[par]}{row}"
                merged_cell = f"{INI_TO_EXCEL[par]}{rowCounter}"
                merged_ws[merged_cell] = input_ws[input_cell].value
                rowCounter += 1

    merged_wb.save(f"{output_dir}/user_data/merged_beta_results.xlsx")


def ParseINIToExcel(fname="_results.ini", update_merge=True):
    """
    Parse data in ini file to excel
    """
    data_ini = configparser.ConfigParser()
    data_ini.read(fname)
    data_ini_section = data_ini.sections()

    # creating new work book and pages
    wb = Workbook()
    wb.create_sheet("DUT")
    wb.create_sheet("TRIG")
    # ws = wb.active

    # look for description file generated by the DAQ
    description_file = None
    for descr in os.listdir("./"):
        if "_Description.ini" in descr:
            description_file = configparser.ConfigParser()
            description_file.read(descr)
            log.info("found DAQ description file")
            break

    try:
        run_description = description_file["Run_Description"]
    except:
        run_description = None
    if run_description:
        try:
            run_number = run_description["Run_Number"]
        except:
            run_number = "NA"
        try:
            dut_name = run_description["DUT_Senor_Name"]
        except:
            dut_name = "DUT_Name_NA"
        try:
            fluence_type = run_description["DUT_Fluence_Type"]
        except:
            fluence_type = "Fluence_Type_NA"
        try:
            fluence = run_description["DUT_Fluence"]
        except:
            fluence = "Fluence_NA"
        try:
            board = run_description["DUT_Readout_Board"]
        except:
            board = "Board_NA"
        try:
            board_number = run_description["DUT_Readout_Board_Number"]
        except:
            board_number = "Board_Number_NA"
        try:
            temperature = run_description["Temperature"]
        except:
            temperature = "Temperature_NA"
        try:
            trig_bias = run_description["Trigger_Voltage"]
        except:
            trig_bias = "Trigger_Voltage_NA"
    else:
        run_number = "NA"
        dut_name = "DUT_Name_NA"
        fluence_type = "Fluence_Type_NA"
        fluence = "Fluence_NA"
        board = "Board_NA"
        board_number = "Board_Number_NA"
        temperature = "Temperature_NA"
        trig_bias = "Trigger_Voltage_NA"

    sensor_name = f"{dut_name}-Fluence {fluence_type}-{fluence}--{board}{board_number}"

    # total transipedence (include amp)
    resistance = 4700

    # start writing data to excel workbook
    for ch in ["DUT", "Trig"]:
        row = 1
        for bias in data_ini_section:
            my_run_num = f"{run_number}->{row}"
            if ch in bias:
                if ch == "DUT":
                    ws = wb["DUT"]
                    run_header = bias.split(",")
                    my_bias = run_header[1]
                    cycle = run_header[2]
                    my_sensor_name = sensor_name
                else:
                    ws = wb["TRIG"]
                    try:
                        my_sensor_name = run_description["Trigger_Sensor_Name"]
                    except:
                        my_sensor_name = "Trigger_Name_NA"
                    run_header = bias.split(",")
                    my_bias = run_header[1].replace("V", "")
                    cycle = run_header[2]
                for par in INI_TO_EXCEL.keys():
                    cell = f"{INI_TO_EXCEL[par]}{row}"
                    if par == "SensorName":
                        ws[cell] = my_sensor_name
                    elif par == "runNumber":
                        ws[cell] = my_run_num
                    elif par == "Temp":
                        ws[cell] = temperature
                    elif par == "Bias":
                        try:
                            ws[cell] = float(my_bias)
                        except:
                            ws[cell] = my_bias
                    elif par == "cycle":
                        ws[cell] = int(cycle)
                    elif par == "Resistance":
                        ws[cell] = float(resistance)
                    else:
                        try:
                            ws[cell] = float(data_ini[bias][par])
                        except:
                            continue
                row += 1
        row += 1  # skip one line

    end_row = len(data_ini_section)

    log.info("Getting time resolution")

    res50_result = Get_Time_Resolution(
        "run_info_v08022018.ini", "50", "keysight", run_number
    )
    res = {}
    res["CFD50Time"] = [item[3] for item in res50_result]
    res["CFD50Time_Err"] = [item[4] for item in res50_result]
    res["cycle"] = [item[5] for item in res50_result]
    res["Bias"] = [item[2] for item in res50_result]
    InjectSortColData(wb["DUT"], 1, "CFD50Time", ["Bias", "cycle"], res)
    InjectSortColData(wb["DUT"], 1, "CFD50Time_Err", ["Bias", "cycle"], res)

    res20_result = Get_Time_Resolution(
        "run_info_v08022018.ini", "20", "keysight", run_number
    )
    res = {}
    res["CFD20Time"] = [item[3] for item in res20_result]
    res["CFD20Time_Err"] = [item[4] for item in res20_result]
    res["cycle"] = [item[5] for item in res20_result]
    res["Bias"] = [item[2] for item in res20_result]
    InjectSortColData(wb["DUT"], 1, "CFD20Time", ["Bias", "cycle"], res)
    InjectSortColData(wb["DUT"], 1, "CFD20Time_Err", ["Bias", "cycle"], res)

    leakage_data = Read_Current("run_info_v08022018.ini")
    leakage_current = {}
    leakage_current["Leakage"] = [item[3] * 1000 for item in leakage_data]
    leakage_current["cycle"] = [item[5] for item in leakage_data]
    leakage_current["Bias"] = [item[2] for item in leakage_data]
    InjectSortColData(wb["DUT"], 1, "Leakage", ["Bias", "cycle"], leakage_current)

    wb.save("_results.xlsx")

    if update_merge:
        MergeExcel("_results.xlsx")


# ===============================================================================
def InjectColData(ws, start_row, par, data_list):
    """
    Method for injecting data
    """
    row = start_row
    for data in data_list:
        cell = f"{par}{row}"
        ws[cell] = data
        row += 1


def InjectSortColData(ws, start_row, par, sort_pars, data_list):
    """
    Similar to InjectColData. data_list is tuple (data, data for sorting)
    """
    sorting_buffer = []
    row = start_row
    for i in range(len(data_list[par])):
        buffer = []
        for sort_par in sort_pars:
            cell = f"{INI_TO_EXCEL[sort_par]}{row}"
            buffer.append(ws[cell].value)
        sorting_buffer.append((row, tuple(buffer)))
        row += 1

    for i, data in enumerate(data_list[par]):
        fill_row = None
        buffer = []
        for name in sort_pars:
            buffer.append(data_list[name][i])

        for item in sorting_buffer:
            if tuple(buffer) == item[1]:
                ws[f"{INI_TO_EXCEL[par]}{item[0]}"] = data


# ===============================================================================
def toROOT():
    import ROOT
    from array import array

    src_wb = None
    output_dir = os.environ["BETASCOPE_SCRIPTS"]
    if os.path.exists("{}/user_data/merged_log.json".format(output_dir)):
        with open("{}/user_data/merged_log.json".format(output_dir)) as f:
            meta_data = json.load(f)
        src_wb = load_workbook(
            "{}/user_data/merged_beta_results.xlsx".format(
                os.environ["BETASCOPE_SCRIPTS"]
            )
        )
    else:
        print("Cannot find merge excel file location")
        return -1

    src_ws = src_wb["DUT"]
    tfile = ROOT.TFile("{}/user_data/merged.root".format(output_dir), "UPDATE")
    tfile.cd()
    for key in meta_data.keys():
        start_row = meta_data[key]["start"]
        end_row = meta_data[key]["end"]
        ttree = ROOT.TTree(key, "from merged excel")
        branches = {}
        for par in par_dict.keys():
            if "SensorName" in par:
                branches[par] = array(
                    "c", str(src_ws[par_dict[par] + str(start_row)].value)
                )
                ttree.Branch(par, branches[par], "{}/C".format(par))
            elif "runNumber" in par:
                continue
            else:
                branches[par] = array("d", [0])
                ttree.Branch(par, branches[par], "{}/D".format(par))
        for row in range(start_row, end_row + 1):
            for par in par_dict.keys():
                if "SensorName" in par or "runNumber" in par:
                    continue
                else:
                    if src_ws[par_dict[par] + str(row)].value == None:
                        src_ws[par_dict[par] + str(row)].value = -99999999999999
                    branches[par][0] = float(src_ws[par_dict[par] + str(row)].value)
            ttree.Fill()
        ttree.Write(key, ROOT.TObject.kOverwrite)
    tfile.Close()


# ===============================================================================
if __name__ == "__main__":
    import argparse

    cml_parser = argparse.ArgumentParser()
    cml_parser.add_argument(
        "-task", dest="task", nargs="?", default="parse", type=str, help="no help!"
    )
    cml_parser.add_argument(
        "-param",
        dest="param",
        nargs="?",
        default="timing",
        type=str,
        help="parameters name!",
    )
    argv = cml_parser.parse_args()

    if argv.task == "parse":
        ParseINIToExcel()
    elif argv.task == "merge":
        MergeExcel()
    elif argv.task == "toROOT":
        toROOT()
    else:
        ParseINIToExcel()